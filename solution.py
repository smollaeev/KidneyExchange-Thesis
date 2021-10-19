from Population import Population
import numpy as np
import copy
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils import get_column_letter

class Solution:
    numberOfGenerations = 15000
    
    def __init__ (self, patients, donors):
        self.patients = patients
        self.donors = donors
        self.bestFitnesses = []
        self.bestChromosomes = []
        self.averageFitnesses = np.zeros (self.numberOfGenerations)

    def GeneticAlgorithm (self):
        pop = Population (0, self.patients, self.donors)
        pop.initialize ()
        
        n = 0
        
        while n < self.numberOfGenerations:     
            pop.calculate_Fitness ()          
            self.bestFitnesses.append (pop.bestFitness)
            self.bestChromosomes.append (pop.bestChromosome)
            self.averageFitnesses [n] = pop.averageFitness
            self.show_BestFitnessOfGeneration (n)

            children=pop.generate_Children ()
            children.mutate ()
            children.chromosomes.extend (copy.deepcopy(pop.elites))
            pop = copy.deepcopy (children)
            pop.generation += 1
            n += 1

    def show_Results (self, i):
        self.save_BestAnswer (i)
        self.show_BestFitnessesDiagram (i)    
        self.show_AverageFitnessesDiagram (i) 
        # for patient in self.patients:
        #     if patient.type == "waitlist":
        #         print (f"{patient.name}: priority==> {patient.priority}")
    
    def show_BestFitnessOfGeneration (self, n):
        print (str (n) + " --> Best Fitness= " + str (self.bestFitnesses [n]))

    def save_BestAnswer (self, repeatedNumber):
        workBook = openpyxl.Workbook ()
        workSheet = workBook.get_sheet_by_name ('Sheet')
        workSheet ['A1'].value = 'Best Fitness'
        workSheet ['B1'].value = self.bestFitnesses [self.numberOfGenerations - 1]

        columnNames = ["Pair", "PatientName", "DonorName", "Compatibility","Priority"]
        for i in range (len (self.bestChromosomes [-1].chains)):
            workSheet = workBook.create_sheet (f'Chain {i + 1}')

            for columnNumber in range (len (columnNames)):
                workSheet [f'{get_column_letter (columnNumber + 1)}1'].value = columnNames [columnNumber]

            for row in range (2, len (self.bestChromosomes [-1].chains [i].nodes) + 2):
                workSheet [f'A{row}'].value = row - 1
                workSheet [f'B{row}'].value = self.bestChromosomes [-1].chains [i].nodes [row - 2].patient.name
                workSheet [f'C{row}'].value = self.bestChromosomes [-1].chains [i].nodes [row - 2].donor.name
                workSheet [f'D{row}'].value = self.bestChromosomes [-1].chains [i].nodes [row - 2].get_Fitness ()
                workSheet [f'E{row}'].value = self.bestChromosomes [-1].chains [i].nodes [row - 2].patient.priority
            # print (str (i+1) + " --> " + str (self.bestChromosomes[-1].pairs[i].patient.name) + "&" + str (self.bestChromosomes [-1].pairs [i].donor.name)
                    # + " --> Compatibility:" + str (self.bestChromosomes [-1].pairs [i].get_Fitness()))
        for i in range (len (self.bestChromosomes [-1].cycles)):
            workSheet = workBook.create_sheet (f'Cycle {i + 1}')

            for columnNumber in range (len (columnNames)):
                workSheet [f'{get_column_letter (columnNumber + 1)}1'].value = columnNames [columnNumber]

            for row in range (2, len (self.bestChromosomes [-1].cycles [i].nodes) + 2):
                workSheet [f'A{row}'].value = row - 1
                workSheet [f'B{row}'].value = self.bestChromosomes [-1].cycles [i].nodes [row - 2].patient.name
                workSheet [f'C{row}'].value = self.bestChromosomes [-1].cycles [i].nodes [row - 2].donor.name
                workSheet [f'D{row}'].value = self.bestChromosomes [-1].cycles [i].nodes [row - 2].get_Fitness ()
        
        workBook.save (f'BestAnswer{repeatedNumber}.xlsx')

    def show_BestFitnessesDiagram (self, i):
        fig = plt.figure ()
        axes = fig.add_axes ([0.1, 0.1, 0.8 , 0.8])
        axes.plot (self.bestFitnesses)
        axes.set_title ('Best Fitness of Answers in Each Generation')
        fig.savefig (f'Best Fitnesses Diagram {i}.jpg')
        # plt.show()

    def show_AverageFitnessesDiagram (self, i):
        fig = plt.figure ()
        axes = fig.add_axes ([0.1, 0.1, 0.8, 0.8])
        axes.plot (self.averageFitnesses)
        axes.set_title ('Average Fitness of Answers in Each Generation')
        fig.savefig (f'Average Fitnesses Diagram {i}.jpg')
        # plt.show()